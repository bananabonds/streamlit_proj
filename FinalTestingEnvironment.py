import random
import re
import time
import datetime
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import Databases.CBLEReviewer_Database2 as data1
import Databases.ARCLectures as data
import urllib.request

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


database_url = "https://raw.githubusercontent.com/bananabonds/streamlit_proj/refs/heads/main/ARCLectureBank.py"

def webDataExtract(chosenTopic):

    try:
        with urllib.request.urlopen(database_url) as response:
            raw_code = response.read().decode("utf-8")

        topicList = {}

        exec(raw_code, topicList)

        chooseTopic = topicList.get(chosenTopic)

    except Exception as e:
        print(f"An error occurred: {e}")

    return chooseTopic


def clean_ocr_text(text):
    """
    Cleans up common OCR / scraping artifacts in question bank text so the
    database doesn't need to be hand-fixed entry by entry:

      1. Rejoins words that were split across a line break with a hyphen
         (e.g. "equilib-\\nrium" -> "equilibrium").
      2. Collapses stray newlines/tabs/runs of whitespace into single
         spaces, while still respecting an intentional blank-line paragraph
         break if one is present in the source.
      3. Inserts a missing space after . , ; : ! ? when the next character
         is a letter jammed directly against the punctuation.
      4. Inserts a missing space where a lowercase word runs directly into
         a following capitalized word with no space between them (a common
         OCR line-wrap artifact), e.g. "the reactionOccurs quickly".
      5. Inserts a missing space where a number runs directly into a
         following capitalized word, e.g. "at 350KThe pressure...".

    Steps 4 and 5 are deliberately conservative (requiring a run of 2+
    lowercase letters on each side) so they don't mangle chemical formulas
    like NaOH, CaCO3, or H2SO4, which have short, alternating-case runs.
    """
    if not isinstance(text, str) or not text:
        return text

    # 1. Dehyphenate words broken across a line wrap.
    text = re.sub(r'(\w)-\s*\n\s*(\w)', r'\1\2', text)

    # 2. Preserve genuine paragraph breaks (blank line), then collapse
    #    every other run of whitespace (including single newlines/tabs)
    #    into one plain space.
    PARAGRAPH_MARKER = "\x00"
    text = re.sub(r'\n\s*\n', PARAGRAPH_MARKER, text)
    text = re.sub(r'\s+', ' ', text)
    text = text.replace(PARAGRAPH_MARKER, '\n\n')

    # 3. Space after punctuation that's glued to the next word.
    text = re.sub(r'([.,;:!?])(?=[A-Za-z])', r'\1 ', text)

    # 4. Space between a lowercase word and a directly-following capitalized
    #    word (word-boundary OCR joins), skipping short/chemical-formula-like runs.
    text = re.sub(r'([a-z]{2,})([A-Z][a-z]{2,})', r'\1 \2', text)

    # 5. Space between a number and a directly-following capitalized word.
    text = re.sub(r'(\d)([A-Z][a-z]{2,})', r'\1 \2', text)

    # 6. Collapse any double spaces the fixes above may have introduced.
    text = re.sub(r' {2,}', ' ', text)

    return text.strip()


def clean_question_bank(questions):
    """Applies clean_ocr_text to every question's text and answer options."""
    cleaned = []
    for q in questions:
        q = dict(q)  # avoid mutating the original source data
        if "text" in q:
            q["text"] = clean_ocr_text(q["text"])
        if "options" in q:
            q["options"] = [clean_ocr_text(opt) for opt in q["options"]]
        cleaned.append(q)
    return cleaned


class CBLESimulator:
    def __init__(self, root):
        self.root = root
        self.root.title("PRC CBLE Simulation - Chemical Engineering Licensure Exam")
        self.root.geometry("1200x750")
        self.root.configure(bg="#F3F4F6")

        # 100 Balanced Board Exam Questions (Day 1: Physical and Chemical Principles

        raw_questions = webDataExtract("Thermo")
        # Auto-clean OCR/scrape artifacts (missing spaces, broken hyphenation,
        # stray line breaks) so the source database doesn't need manual fixing.
        cleaned_questions = clean_question_bank(list(raw_questions))
        # CRITICAL: Shuffle data randomly at startup execution
        self.questions = cleaned_questions
        random.shuffle(self.questions)
        
        # State Tracking
        self.current_index = 0
        self.total_questions = len(self.questions)
        self.user_answers = {i: None for i in range(self.total_questions)}
        self.flagged_questions = {i: False for i in range(self.total_questions)}
        self.time_left = 3 * 3600  # 3 Hours

        # NEW: Per-question time tracking (seconds spent on each question)
        self.question_times = {i: 0.0 for i in range(self.total_questions)}
        self.question_start_time = None  # set whenever a question is loaded/viewed

        self.selected_option = tk.IntVar(value=-1)

        self.create_widgets()
        self.update_timer()
        self.load_question()

    def create_widgets(self):
        # 1. Top Bar (Header & Timer)
        top_bar = tk.Frame(self.root, bg="#1E3A8A", height=60)
        top_bar.pack(fill=tk.X, side=tk.TOP)
        
        title_lbl = tk.Label(top_bar, text="BOARD OF CHEMICAL ENGINEERING - CBLE RANDOMIZED SIMULATION (DAY 1)", fg="white", bg="#1E3A8A", font=("Arial", 12, "bold"))
        title_lbl.pack(side=tk.LEFT, padx=15, pady=15)
        
        self.timer_lbl = tk.Label(top_bar, text="Time Remaining: 03:00:00", fg="#EF4444", bg="#1E3A8A", font=("Arial", 13, "bold"))
        self.timer_lbl.pack(side=tk.RIGHT, padx=15, pady=15)

        # 2. Main Body Container
        main_body = tk.Frame(self.root, bg="#F3F4F6")
        main_body.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left Column: Exam Presentation
        left_col = tk.Frame(main_body, bg="#F3F4F6")
        left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        # Question Panel
        q_frame = tk.LabelFrame(left_col, text=" Question Box ", font=("Arial", 10, "bold"), bg="white", bd=2, relief=tk.GROOVE)
        q_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.q_text = tk.Text(q_frame, font=("Arial", 11), wrap=tk.WORD, bg="white", bd=0, highlightthickness=0)
        self.q_text.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        self.q_text.config(state=tk.DISABLED)

        # Multiple Choice Options Frame
        self.options_frame = tk.LabelFrame(left_col, text=" Select Your Answer ", font=("Arial", 10, "bold"), bg="white", bd=2, relief=tk.GROOVE)
        self.options_frame.pack(fill=tk.X, pady=5)

        self.radio_buttons = []
        for i in range(4):
            rb = tk.Radiobutton(self.options_frame, text="", variable=self.selected_option, value=i, font=("Arial", 11), bg="white", anchor="w", command=self.save_answer, padx=20, pady=8)
            rb.pack(fill=tk.X)
            self.radio_buttons.append(rb)

        # Bottom Actions Panel
        action_frame = tk.Frame(left_col, bg="#F3F4F6")
        action_frame.pack(fill=tk.X, pady=10)

        self.btn_prev = tk.Button(action_frame, text="◀ Previous", font=("Arial", 10, "bold"), width=12, command=self.prev_question)
        self.btn_prev.pack(side=tk.LEFT)

        self.btn_flag = tk.Button(action_frame, text="🏳 Flag Question", font=("Arial", 10, "bold"), fg="#B45309", bg="#FEF3C7", width=15, command=self.toggle_flag)
        self.btn_flag.pack(side=tk.LEFT, padx=10)

        self.btn_next = tk.Button(action_frame, text="Next ▶", font=("Arial", 10, "bold"), width=12, command=self.next_question)
        self.btn_next.pack(side=tk.LEFT)

        self.btn_submit = tk.Button(action_frame, text="Submit Exam", font=("Arial", 10, "bold"), fg="white", bg="#DC2626", width=12, command=self.confirm_submission)
        self.btn_submit.pack(side=tk.RIGHT)

        # Live Summary Notice Status Bar
        self.status_lbl = tk.Label(left_col, text="Answered: 0 | Flagged: 0 | Remaining: 100", font=("Arial", 10, "italic"), bg="#F3F4F6", fg="#4B5563")
        self.status_lbl.pack(side=tk.BOTTOM, fill=tk.X, pady=5)

        # Right Column: Question Navigation Grid Matrix
        right_col = tk.LabelFrame(main_body, text=" Question Matrix ", font=("Arial", 10, "bold"), bg="white", width=280, bd=2, relief=tk.GROOVE)
        right_col.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))
        right_col.pack_propagate(False)

        self.matrix_canvas = tk.Canvas(right_col, bg="white", borderwidth=0, highlightthickness=0)
        matrix_scrollbar = ttk.Scrollbar(right_col, orient="vertical", command=self.matrix_canvas.yview)
        self.matrix_container = tk.Frame(self.matrix_canvas, bg="white")

        self._matrix_window = self.matrix_canvas.create_window((0, 0), window=self.matrix_container, anchor="nw")

        self.matrix_container.bind(
            "<Configure>",
            lambda e: self.matrix_canvas.configure(scrollregion=self.matrix_canvas.bbox("all"))
        )
        # Keep the inner frame's width in sync with the canvas so the button
        # grid re-wraps correctly if the panel is ever resized.
        self.matrix_canvas.bind(
            "<Configure>",
            lambda e: self.matrix_canvas.itemconfigure(self._matrix_window, width=e.width)
        )
        self.matrix_canvas.configure(yscrollcommand=matrix_scrollbar.set)

        self.matrix_canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        matrix_scrollbar.pack(side="right", fill="y")

        # Mouse-wheel scrolling for the matrix panel. Bound only while the
        # cursor is over the panel so it doesn't hijack scrolling elsewhere
        # in the window. Handles Windows/Mac (<MouseWheel>) and Linux
        # (<Button-4>/<Button-5>) separately.
        def _on_matrix_mousewheel(event):
            if event.num == 4:          # Linux scroll up
                delta = -1
            elif event.num == 5:        # Linux scroll down
                delta = 1
            else:                       # Windows / Mac: event.delta is +/-120 (or +/-1 on Mac)
                delta = -1 if event.delta > 0 else 1
            self.matrix_canvas.yview_scroll(delta, "units")

        def _bind_matrix_mousewheel(event):
            self.matrix_canvas.bind_all("<MouseWheel>", _on_matrix_mousewheel)
            self.matrix_canvas.bind_all("<Button-4>", _on_matrix_mousewheel)
            self.matrix_canvas.bind_all("<Button-5>", _on_matrix_mousewheel)

        def _unbind_matrix_mousewheel(event):
            self.matrix_canvas.unbind_all("<MouseWheel>")
            self.matrix_canvas.unbind_all("<Button-4>")
            self.matrix_canvas.unbind_all("<Button-5>")

        self.matrix_canvas.bind("<Enter>", _bind_matrix_mousewheel)
        self.matrix_canvas.bind("<Leave>", _unbind_matrix_mousewheel)

        self.matrix_buttons = {}
        for idx in range(self.total_questions):
            row = idx // 5   
            col = idx % 5
            btn = tk.Button(self.matrix_container, text=str(idx + 1), font=("Arial", 8, "bold"), width=4, height=2, bg="#E5E7EB", relief=tk.FLAT, command=lambda i=idx: self.jump_to_question(i))
            btn.grid(row=row, column=col, padx=3, pady=3)
            self.matrix_buttons[idx] = btn

    # ------------------------------------------------------------------
    # NEW: centralized time tracking helper.
    # Call this BEFORE moving away from the currently displayed question
    # so the elapsed seconds get credited to the right question index.
    # ------------------------------------------------------------------
    def _bank_elapsed_time(self):
        if self.question_start_time is not None:
            elapsed = time.time() - self.question_start_time
            self.question_times[self.current_index] += elapsed
        self.question_start_time = None

    def load_question(self):
        q = self.questions[self.current_index]
        
        self.q_text.config(state=tk.NORMAL)
        self.q_text.delete("1.0", tk.END)
        # Formats the current sequence string smoothly dynamically
        self.q_text.insert(tk.END, f"Question {self.current_index + 1}:\n" + q["text"])
        self.q_text.config(state=tk.DISABLED)
        
        for i, opt in enumerate(q["options"]):
            self.radio_buttons[i].config(text=opt)
            
        saved = self.user_answers[self.current_index]
        self.selected_option.set(saved if saved is not None else -1)
        
        if self.flagged_questions[self.current_index]:
            self.btn_flag.config(text="🏳 Unflag", bg="#F59E0B", fg="white")
        else:
            self.btn_flag.config(text="🏳 Flag Question", bg="#FEF3C7", fg="#B45309")

        self.refresh_matrix()

        # Start (or restart) the clock for whichever question is now on screen
        self.question_start_time = time.time()

    def save_answer(self):
        self.user_answers[self.current_index] = self.selected_option.get()
        self.refresh_matrix()

    def toggle_flag(self):
        self.flagged_questions[self.current_index] = not self.flagged_questions[self.current_index]
        if self.flagged_questions[self.current_index]:
            self.btn_flag.config(text="🏳 Unflag", bg="#F59E0B", fg="white")
        else:
            self.btn_flag.config(text="🏳 Flag Question", bg="#FEF3C7", fg="#B45309")
        self.refresh_matrix()

    def refresh_matrix(self):
        answered_count = 0
        flagged_count = 0

        for idx, btn in self.matrix_buttons.items():
            is_answered = self.user_answers[idx] is not None
            is_flagged = self.flagged_questions[idx]

            if is_answered: answered_count += 1
            if is_flagged: flagged_count += 1

            if idx == self.current_index:
                btn.config(bd=1, relief=tk.SOLID, highlightbackground="black")
            else:
                btn.config(bd=0, relief=tk.FLAT)

            if is_flagged:
                btn.config(bg="#F59E0B", fg="white")
            elif is_answered:
                btn.config(bg="#10B981", fg="white")
            else:
                btn.config(bg="#E5E7EB", fg="black")

        remaining = self.total_questions - answered_count
        self.status_lbl.config(text=f"Answered: {answered_count} | Flagged: {flagged_count} | Unanswered Remaining: {remaining}")

        self.btn_prev.config(state=tk.DISABLED if self.current_index == 0 else tk.NORMAL)
        self.btn_next.config(state=tk.DISABLED if self.current_index == self.total_questions - 1 else tk.NORMAL)

    def jump_to_question(self, index):
        self._bank_elapsed_time()
        self.current_index = index
        self.load_question()

    def next_question(self):
        if self.current_index < self.total_questions - 1:
            self._bank_elapsed_time()
            self.current_index += 1
            self.load_question()

    def prev_question(self):
        if self.current_index > 0:
            self._bank_elapsed_time()
            self.current_index -= 1
            self.load_question()

    def update_timer(self):
        if self.time_left > 0:
            self.time_left -= 1
            hrs = self.time_left // 3600
            mins = (self.time_left % 3600) // 60
            secs = self.time_left % 60
            self.timer_lbl.config(text=f"Time Remaining: {hrs:02d}:{mins:02d}:{secs:02d}")
            self.root.after(1000, self.update_timer)
        else:
            self.show_report()

    def confirm_submission(self):
        unanswered = sum(1 for ans in self.user_answers.values() if ans is None)
        msg = "Are you sure you want to submit and end the exam?"
        if unanswered > 0:
            msg = f"You still have {unanswered} unanswered question(s). \n\n" + msg
            
        if messagebox.askyesno("Confirm Submit", msg):
            self.show_report()

    @staticmethod
    def _format_duration(seconds):
        seconds = int(round(seconds))
        hrs = seconds // 3600
        mins = (seconds % 3600) // 60
        secs = seconds % 60
        if hrs > 0:
            return f"{hrs:02d}:{mins:02d}:{secs:02d}"
        return f"{mins:02d}:{secs:02d}"

    def show_report(self):
        # Make sure whatever question was on-screen at submit time gets its
        # elapsed time banked before we build the report.
        self._bank_elapsed_time()

        self.score = 0
        self.all_items_report = []
        self.sort_by_wrong_first = False

        for i, q in enumerate(self.questions):
            user_ans = self.user_answers[i]
            correct_ans = q["correct"]
            is_correct = (user_ans == correct_ans)
            
            if is_correct:
                self.score += 1

            if user_ans is not None and 0 <= user_ans < len(q["options"]):
                chosen_text = q["options"][user_ans]
            else:
                chosen_text = "No Answer Provided (Skipped)"

            self.all_items_report.append({
                "number": i + 1,
                "is_correct": is_correct,
                "is_flagged": self.flagged_questions[i],
                "time_spent_seconds": self.question_times[i],
                "question": q["text"],
                "your_answer": chosen_text,
                "correct_answer": q["options"][correct_ans]
            })

        self.report_win = tk.Toplevel(self.root)
        self.report_win.title("ChELE Diagnostic Correction Report")
        self.report_win.geometry("900x700")
        self.report_win.configure(bg="white")
        self.report_win.grab_set()

        summary_frame = tk.Frame(self.report_win, bg="#1E3A8A")
        summary_frame.pack(fill=tk.X, padx=15, pady=10)
        
        status_text = "PASSED" if (self.score / self.total_questions) >= 0.70 else "FAILED"
        header_lbl = tk.Label(summary_frame, text=f"EXAM RESULTS: {self.score}/{self.total_questions} ({status_text})", fg="white", bg="#1E3A8A", font=("Arial", 14, "bold"))
        header_lbl.pack(pady=10)

        control_bar = tk.Frame(self.report_win, bg="#E5E7EB")
        control_bar.pack(fill=tk.X, padx=15, pady=(5, 0))
        
        self.btn_sort = tk.Button(control_bar, text="Sort Order: Wrong Answers First ↕", font=("Arial", 10, "bold"), bg="#FFFFFF", fg="#1E3A8A", command=self.toggle_report_sort)
        self.btn_sort.pack(side=tk.LEFT, padx=10, pady=5)

        self.btn_copy = tk.Button(control_bar, text="📋 Copy Report to Clipboard", font=("Arial", 10, "bold"), bg="#FFFFFF", fg="#1E3A8A", command=self.copy_report_to_clipboard)
        self.btn_copy.pack(side=tk.LEFT, padx=10, pady=5)

        self.btn_export = tk.Button(control_bar, text="📊 Export to Excel", font=("Arial", 10, "bold"), bg="#FFFFFF", fg="#166534", command=self.export_report_to_excel)
        self.btn_export.pack(side=tk.LEFT, padx=10, pady=5)

        # Single scrollable, selectable, copy-pasteable Text widget replaces
        # the old per-question card layout.
        text_frame = tk.Frame(self.report_win, bg="white")
        text_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)

        self.report_text = tk.Text(text_frame, font=("Consolas", 10), wrap=tk.WORD, bg="#FAFAFA", fg="#111827", bd=1, relief=tk.SOLID)
        report_scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=self.report_text.yview)
        self.report_text.configure(yscrollcommand=report_scrollbar.set)

        self.report_text.pack(side="left", fill="both", expand=True)
        report_scrollbar.pack(side="right", fill="y")

        self.render_report_text()

        def exit_all():
            self.report_win.destroy()
            self.root.destroy()

        close_btn = tk.Button(self.report_win, text="Close and Finish Simulator", font=("Arial", 11, "bold"), bg="#4B5563", fg="white", command=exit_all)
        close_btn.pack(side=tk.BOTTOM, fill=tk.X, padx=15, pady=10)

    def _build_report_string(self):
        """Builds the full plain-text, copy-pasteable exam report."""
        total_time_used = sum(self.question_times.values())
        pct = (self.score / self.total_questions) * 100 if self.total_questions else 0
        status_text = "PASSED" if (self.score / self.total_questions) >= 0.70 else "FAILED"

        if self.sort_by_wrong_first:
            display_list = sorted(self.all_items_report, key=lambda x: x["is_correct"])
            sort_label = "Wrong Answers First"
        else:
            display_list = sorted(self.all_items_report, key=lambda x: x["number"])
            sort_label = "Sequential (1 - N)"

        lines = []
        sep = "=" * 78
        subsep = "-" * 78

        lines.append(sep)
        lines.append("PRC CBLE SIMULATION - EXAM DIAGNOSTIC REPORT")
        lines.append(sep)
        lines.append(f"Generated:      {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Score:          {self.score}/{self.total_questions} ({pct:.1f}%) - {status_text}")
        lines.append(f"Total Time Used:{' ':1}{self._format_duration(total_time_used)}")
        lines.append(f"Sort Order:     {sort_label}")
        lines.append(sep)
        lines.append("")

        for item in display_list:
            status_tag = "CORRECT" if item["is_correct"] else "INCORRECT/SKIPPED"
            flag_tag = "FLAGGED" if item["is_flagged"] else "Not Flagged"
            time_tag = self._format_duration(item["time_spent_seconds"])

            lines.append(subsep)
            lines.append(f"Q{item['number']}  |  {status_tag}  |  {flag_tag}  |  Time Spent: {time_tag}")
            lines.append(subsep)
            lines.append(f"Question: {item['question']}")
            lines.append(f"Your Answer:    {item['your_answer']}")
            lines.append(f"Correct Answer: {item['correct_answer']}")
            lines.append("")

        return "\n".join(lines)

    def render_report_text(self):
        report_str = self._build_report_string()
        self.report_text.config(state=tk.NORMAL)
        self.report_text.delete("1.0", tk.END)
        self.report_text.insert(tk.END, report_str)
        # Leave state NORMAL so the user can select/copy freely; text
        # is regenerated wholesale on sort-toggle so accidental edits
        # by the user don't persist across a re-render.

    def copy_report_to_clipboard(self):
        report_str = self._build_report_string()
        self.root.clipboard_clear()
        self.root.clipboard_append(report_str)
        self.root.update()  # ensures clipboard content persists after app focus changes
        messagebox.showinfo("Copied", "The full report has been copied to your clipboard.")

    def export_report_to_excel(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror(
                "Missing Dependency",
                "Excel export requires the 'openpyxl' package.\n\nInstall it with:\n    pip install openpyxl"
            )
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"CBLE_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            title="Save Exam Report As"
        )
        if not file_path:
            return

        wb = openpyxl.Workbook()

        # --- Summary sheet ---
        ws_summary = wb.active
        ws_summary.title = "Summary"
        total_time_used = sum(self.question_times.values())
        pct = (self.score / self.total_questions) * 100 if self.total_questions else 0
        status_text = "PASSED" if (self.score / self.total_questions) >= 0.70 else "FAILED"

        summary_rows = [
            ("Generated", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Score", f"{self.score}/{self.total_questions}"),
            ("Percentage", f"{pct:.1f}%"),
            ("Result", status_text),
            ("Total Time Used", self._format_duration(total_time_used)),
        ]
        for r_idx, (label, value) in enumerate(summary_rows, start=1):
            ws_summary.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=r_idx, column=2, value=value)
        ws_summary.column_dimensions["A"].width = 20
        ws_summary.column_dimensions["B"].width = 30

        # --- Detail sheet ---
        ws = wb.create_sheet("Question Detail")
        headers = ["Q#", "Status", "Flagged", "Time Spent (mm:ss)", "Time Spent (sec)",
                   "Question", "Your Answer", "Correct Answer"]
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sorted_for_export = sorted(self.all_items_report, key=lambda x: x["number"])
        correct_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
        incorrect_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

        for row_idx, item in enumerate(sorted_for_export, start=2):
            row_fill = correct_fill if item["is_correct"] else incorrect_fill
            values = [
                item["number"],
                "Correct" if item["is_correct"] else "Incorrect/Skipped",
                "Yes" if item["is_flagged"] else "No",
                self._format_duration(item["time_spent_seconds"]),
                round(item["time_spent_seconds"], 1),
                item["question"],
                item["your_answer"],
                item["correct_answer"],
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="top", wrap_text=(col_idx >= 6))

        col_widths = [6, 18, 10, 18, 16, 60, 40, 40]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        try:
            wb.save(file_path)
            messagebox.showinfo("Export Successful", f"Report exported to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not save the file:\n{e}")

    def toggle_report_sort(self):
        self.sort_by_wrong_first = not self.sort_by_wrong_first
        if self.sort_by_wrong_first:
            self.btn_sort.config(text="Sort Order: Sequential Chronological (1-100) ↕")
        else:
            self.btn_sort.config(text="Sort Order: Wrong Answers First ↕")
        self.render_report_text()


if __name__ == "__main__":
    root = tk.Tk()
    app = CBLESimulator(root)
    root.mainloop()